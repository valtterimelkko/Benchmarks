from __future__ import annotations

import io
import json
import re
import time
from datetime import datetime, timezone
from typing import Any, Callable

import openpyxl
import pandas as pd
import requests
from bs4 import BeautifulSoup

from .models import BenchmarkRow, BenchmarkSnapshot

USER_AGENT = "letsautomate-benchmark-dashboard/0.1 (+private weekly dashboard)"
TIMEOUT = 35


ORG_PATTERNS = [
    ("openai", "OpenAI"),
    ("gpt", "OpenAI"),
    ("claude", "Anthropic"),
    ("anthropic", "Anthropic"),
    ("gemini", "Google"),
    ("google", "Google"),
    ("kimi", "Moonshot AI"),
    ("moonshot", "Moonshot AI"),
    ("glm", "Z.AI"),
    ("z.ai", "Z.AI"),
    ("deepseek", "DeepSeek"),
    ("qwen", "Alibaba"),
    ("alibaba", "Alibaba"),
    ("minimax", "MiniMax"),
    ("mimo", "Xiaomi"),
    ("mistral", "Mistral"),
    ("xai", "xAI"),
    ("grok", "xAI"),
    ("meta", "Meta"),
    ("llama", "Meta"),
]


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


_FETCH_RETRY_DELAYS: tuple[float, ...] = (5.0, 15.0)


def _find_col(headers: list[str], *candidates: str) -> int | None:
    """Return column index for first matching candidate (exact then substring)."""
    lh = [h.lower().strip() for h in headers]
    for candidate in candidates:
        c = candidate.lower()
        if c in lh:
            return lh.index(c)
    for candidate in candidates:
        c = candidate.lower()
        for i, h in enumerate(lh):
            if c in h:
                return i
    return None


def _fetch_with_retry(
    fetcher: Callable[[], list[BenchmarkRow]],
    delays: tuple[float, ...] = _FETCH_RETRY_DELAYS,
) -> list[BenchmarkRow]:
    """Call fetcher up to len(delays)+1 times with exponential backoff on failure."""
    last_exc: Exception | None = None
    for attempt in range(len(delays) + 1):
        try:
            rows = fetcher()
            if not rows:
                raise ValueError("No parseable leaderboard rows found")
            return rows
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if attempt < len(delays):
                time.sleep(delays[attempt])
    raise last_exc  # type: ignore[misc]


def _validate_snapshot(
    rows: list[BenchmarkRow],
    min_rows: int,
    score_range: tuple[float, float] | None,
) -> str | None:
    """Return a warning string if the snapshot looks suspicious, else None."""
    issues: list[str] = []
    if len(rows) < min_rows:
        issues.append(f"Only {len(rows)} row(s) returned — expected ≥{min_rows}.")
    if score_range is not None and rows:
        lo, hi = score_range
        top = rows[0].score
        if not lo <= top <= hi:
            issues.append(f"Top score {top} is outside expected range {lo}–{hi} — possible parser or source issue.")
    return " ".join(issues) if issues else None


def infer_org(model: str | None) -> str | None:
    if not model:
        return None
    lowered = model.lower()
    for needle, org in ORG_PATTERNS:
        if needle in lowered:
            return org
    return None


def normalise_org(org: str | None) -> str | None:
    if not org:
        return None
    lookup = {"anthropic": "Anthropic", "openai": "OpenAI", "google": "Google", "deepseek": "DeepSeek", "moonshot ai": "Moonshot AI", "moonshot": "Moonshot AI", "z.ai": "Z.AI", "alibaba": "Alibaba"}
    return lookup.get(org.strip().lower(), org.strip())


def clean_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split())


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", ""))
    return float(match.group(0)) if match else None


def fetch_text(url: str) -> str:
    response = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    response.raise_for_status()
    return response.text


def fetch_bytes(url: str) -> bytes:
    response = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    response.raise_for_status()
    return response.content


def parse_deepswe_html(html: str) -> list[BenchmarkRow]:
    pattern = re.compile(
        r'\{model:"(?P<model>[^"]+)",harness:"(?P<harness>[^"]+)"(?:,reasoning_effort:"(?P<effort>[^"]*)")?.*?'
        r'pass_rate:(?P<pass_rate>\d+(?:\.\d+)?).*?'
        r'n_tasks_attempted:(?P<attempted>\d+).*?'
        r'n_tasks_passed_any:(?P<passed>\d+).*?'
        r'mean_cost_usd:(?P<cost>\d+(?:\.\d+)?)',
        re.DOTALL,
    )
    rows: list[BenchmarkRow] = []
    for match in pattern.finditer(html):
        model = match.group("model")
        effort = match.group("effort") or ""
        score = round(float(match.group("pass_rate")) * 100, 2)
        model_label = f"{model} [{effort}]" if effort else model
        rows.append(
            BenchmarkRow(
                rank=0,
                model=model_label,
                organization=infer_org(model),
                score=score,
                score_unit="% solve rate",
                metadata={
                    "harness": match.group("harness"),
                    "reasoning_effort": effort or None,
                    "tasks_attempted": int(match.group("attempted")),
                    "tasks_passed_any": int(match.group("passed")),
                    "mean_cost_usd": round(float(match.group("cost")), 4),
                },
            )
        )
    rows.sort(key=lambda row: row.score, reverse=True)
    for index, row in enumerate(rows, start=1):
        row.rank = index
    return rows


def parse_terminal_html(html: str) -> list[BenchmarkRow]:
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if table is None:
        return []
    all_trs = table.find_all("tr")
    header_tr = next((tr for tr in all_trs if tr.find("th")), None)
    if header_tr is None:
        return []
    headers = [clean_text(th.get_text()) for th in header_tr.find_all(["th", "td"])]
    rank_col = _find_col(headers, "rank")
    model_col = _find_col(headers, "model")
    agent_col = _find_col(headers, "agent")
    date_col = _find_col(headers, "date")
    agent_org_col = _find_col(headers, "agent org")
    model_org_col = _find_col(headers, "model org")
    accuracy_col = _find_col(headers, "accuracy")
    if rank_col is None or model_col is None or accuracy_col is None:
        return []
    rows: list[BenchmarkRow] = []
    for tr in all_trs:
        if tr is header_tr:
            continue
        cells = [clean_text(td.get_text(" ", strip=True)) for td in tr.find_all("td")]
        required_idx = max(c for c in [rank_col, model_col, accuracy_col])
        if len(cells) <= required_idx:
            continue
        rank_val = cells[rank_col]
        if not rank_val.isdigit():
            continue
        score = as_float(cells[accuracy_col])
        if score is None:
            continue
        model_val = cells[model_col]
        rows.append(
            BenchmarkRow(
                rank=int(rank_val),
                model=model_val,
                organization=normalise_org(cells[model_org_col] if model_org_col is not None and model_org_col < len(cells) else "") or infer_org(model_val),
                score=score,
                score_unit="% accuracy",
                date=cells[date_col] if date_col is not None and date_col < len(cells) else None,
                metadata={
                    "agent": cells[agent_col] if agent_col is not None and agent_col < len(cells) else "",
                    "agent_org": cells[agent_org_col] if agent_org_col is not None and agent_org_col < len(cells) else "",
                    "raw_accuracy": cells[accuracy_col],
                },
            )
        )
    return rows


def parse_benchlm_next_data(html: str, *, score_unit: str = "% score") -> list[BenchmarkRow]:
    match = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>', html, re.DOTALL)
    if not match:
        return []
    data = json.loads(match.group(1))
    leaderboard = data.get("props", {}).get("pageProps", {}).get("leaderboard", [])
    rows: list[BenchmarkRow] = []
    for index, item in enumerate(leaderboard, start=1):
        score = as_float(item.get("score"))
        if score is None:
            continue
        rows.append(
            BenchmarkRow(
                rank=index,
                model=clean_text(item.get("model")),
                organization=normalise_org(item.get("creator")) or infer_org(item.get("model")),
                score=round(score, 2),
                score_unit=score_unit,
                metadata={
                    "source_type": item.get("sourceType"),
                    "context_window": item.get("contextWindow"),
                    "overall_score": item.get("overallScore"),
                    "benchlm_slug": item.get("slug"),
                },
            )
        )
    return rows


def parse_osworld_workbook(content: bytes) -> list[BenchmarkRow]:
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook["Eval Results"] if "Eval Results" in workbook.sheetnames else workbook.active
    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {name: i for i, name in enumerate(headers)}
    rows: list[BenchmarkRow] = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        model = clean_text(raw[index.get("Model", 0)])
        score = as_float(raw[index.get("Success rate", 10)])
        if not model or score is None:
            continue
        date_value = raw[index.get("Date", 9)] if "Date" in index else None
        date = date_value.strftime("%Y-%m-%d") if hasattr(date_value, "strftime") else (clean_text(date_value) or None)
        max_steps_idx = index.get("Max steps")
        approach_idx = index.get("Approach type")
        success_idx = index.get("Success/Total")
        rows.append(
            BenchmarkRow(
                rank=0,
                model=model,
                organization=normalise_org(clean_text(raw[index.get("Institution", 1)])) or infer_org(model),
                score=round(score, 2),
                score_unit="% success rate",
                date=date,
                metadata={
                    "approach_type": clean_text(raw[approach_idx]) if approach_idx is not None else None,
                    "max_steps": raw[max_steps_idx] if max_steps_idx is not None else None,
                    "success_total": clean_text(raw[success_idx]) if success_idx is not None else None,
                },
            )
        )
    rows.sort(key=lambda row: row.score, reverse=True)
    for index, row in enumerate(rows, start=1):
        row.rank = index
    return rows


def parse_longbench_html(html: str) -> list[BenchmarkRow]:
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if table is None:
        return []
    all_trs = table.find_all("tr")
    header_tr = next((tr for tr in all_trs if tr.find("th")), None)
    if header_tr is None:
        return []
    headers = [clean_text(th.get_text()) for th in header_tr.find_all(["th", "td"])]
    model_col = _find_col(headers, "model")
    params_col = _find_col(headers, "params")
    context_col = _find_col(headers, "context")
    date_col = _find_col(headers, "date")
    overall_col = _find_col(headers, "overall")
    cot_col = _find_col(headers, "w/ cot", "cot")
    if model_col is None:
        return []
    rows: list[BenchmarkRow] = []
    for tr in all_trs:
        if tr is header_tr:
            continue
        cells = [clean_text(td.get_text(" ", strip=True)) for td in tr.find_all("td")]
        if model_col >= len(cells):
            continue
        model_cell = cells[model_col]
        if not model_cell or model_cell.lower() in ("model", "human"):
            continue
        score: float | None = None
        if overall_col is not None and overall_col < len(cells) and cells[overall_col] not in ("-", ""):
            score = as_float(cells[overall_col])
        if score is None and cot_col is not None and cot_col < len(cells):
            score = as_float(cells[cot_col])
        if score is None:
            continue
        org = infer_org(model_cell)
        for candidate in ["Google", "Alibaba", "DeepSeek", "OpenAI", "Anthropic", "Meta", "Moonshot AI", "Z.AI", "Mistral"]:
            if candidate.lower() in model_cell.lower():
                org = candidate
                break
        model = model_cell.replace("🧠", "").strip()
        if org and model.endswith(org):
            model = model[: -len(org)].strip()
        rows.append(
            BenchmarkRow(
                rank=0,
                model=model,
                organization=org,
                score=round(score, 2),
                score_unit="% overall",
                date=cells[date_col] if date_col is not None and date_col < len(cells) else None,
                metadata={
                    "params": cells[params_col] if params_col is not None and params_col < len(cells) else None,
                    "context": cells[context_col] if context_col is not None and context_col < len(cells) else None,
                },
            )
        )
    rows.sort(key=lambda row: row.score, reverse=True)
    for index, row in enumerate(rows, start=1):
        row.rank = index
    return rows


def parse_aa_gdpval_table(html: str) -> list[BenchmarkRow]:
    """Parse the Artificial Analysis GDPval-AA Elo table.

    Expected columns: (blank rank) | Creator | Name | Elo | CI | Release Date
    Negative Elo values use U+2212 MINUS SIGN (e.g. '−16'), not a plain hyphen.
    """
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if table is None:
        return []
    rows: list[BenchmarkRow] = []
    for tr in table.find_all("tr")[1:]:  # skip header row
        cells = [clean_text(td.get_text(" ", strip=True)) for td in tr.find_all("td")]
        if len(cells) < 4:
            continue
        rank_str = cells[0]
        if not rank_str.isdigit():
            continue
        creator, name = cells[1], cells[2]
        score = as_float(cells[3].replace("−", "-"))  # normalise typographic minus
        if score is None:
            continue
        ci_str = cells[4] if len(cells) > 4 else ""
        date_str = cells[5] if len(cells) > 5 else None
        ci_parts = re.findall(r"[+-]?\d+", ci_str)
        rows.append(
            BenchmarkRow(
                rank=int(rank_str),
                model=name,
                organization=normalise_org(creator) or infer_org(name),
                score=round(score, 1),
                score_unit="Elo rating",
                date=date_str or None,
                metadata={
                    "ci_lower": int(ci_parts[0]) if len(ci_parts) >= 1 else None,
                    "ci_upper": int(ci_parts[1]) if len(ci_parts) >= 2 else None,
                },
            )
        )
    return rows


def parse_aa_intelligence_index_ldjson(html: str) -> list[BenchmarkRow]:
    """Parse the Artificial Analysis Intelligence Index from schema.org JSON-LD blocks.

    The AA homepage embeds multiple <script type="application/ld+json"> Dataset
    blocks. The headline composite is named 'Artificial Analysis Intelligence
    Index' (score key 'intelligenceIndex'); a shorter 'Intelligence' dataset
    (score key 'artificialAnalysisIntelligenceIndex') mirrors the same ranking
    and acts as a fallback if AA renames the headline block. Names must match
    exactly — e.g. 'Intelligence Index by Open Weights / Proprietary' is a
    different cut of the data and must not be picked up.
    """
    candidates = {
        "Artificial Analysis Intelligence Index": "intelligenceIndex",
        "Intelligence": "artificialAnalysisIntelligenceIndex",
    }
    found: dict[str, list[BenchmarkRow]] = {}
    for match in re.finditer(r'<script type="application/ld\+json">(.*?)</script>', html, re.DOTALL):
        try:
            dataset = json.loads(match.group(1))
        except (json.JSONDecodeError, ValueError):
            continue
        if not isinstance(dataset, dict) or dataset.get("@type") != "Dataset":
            continue
        name = str(dataset.get("name") or "").strip()
        score_key = candidates.get(name)
        if score_key is None or name in found:
            continue
        data = dataset.get("data")
        if not isinstance(data, list):
            continue
        description = str(dataset.get("description") or "")
        version_match = re.search(r"Intelligence Index\s+(v[\d.]+)", description)
        index_version = version_match.group(1) if version_match else None
        rows: list[BenchmarkRow] = []
        for item in data:
            if not isinstance(item, dict):
                continue
            label = clean_text(item.get("label"))
            score = as_float(item.get(score_key))
            if not label or score is None:
                continue
            details = clean_text(item.get("detailsUrl")) or None
            if details and details.startswith("/"):
                details = f"https://artificialanalysis.ai{details}"
            rows.append(
                BenchmarkRow(
                    rank=0,
                    model=label,
                    organization=infer_org(label),
                    score=round(score, 1),
                    score_unit="index points",
                    metadata={
                        "details_url": details,
                        "index_version": index_version,
                    },
                )
            )
        if rows:
            found[name] = rows
    # Prefer the full headline dataset; the shorter 'Intelligence' block is the fallback.
    rows = found.get("Artificial Analysis Intelligence Index") or found.get("Intelligence") or []
    rows.sort(key=lambda row: row.score, reverse=True)
    for index, row in enumerate(rows, start=1):
        row.rank = index
    return rows


def parse_lmarena_rows(payload: dict[str, Any]) -> list[BenchmarkRow]:
    rows: list[BenchmarkRow] = []
    for wrapper in payload.get("rows", []):
        item = wrapper.get("row", wrapper)
        score = as_float(item.get("rating"))
        if score is None:
            continue
        rows.append(
            BenchmarkRow(
                rank=int(as_float(item.get("rank")) or len(rows) + 1),
                model=clean_text(item.get("model_name")),
                organization=normalise_org(clean_text(item.get("organization"))) or infer_org(item.get("model_name")),
                score=round(score, 2),
                score_unit="Arena rating",
                date=item.get("leaderboard_publish_date"),
                metadata={
                    "rating_lower": item.get("rating_lower"),
                    "rating_upper": item.get("rating_upper"),
                    "votes": item.get("vote_count"),
                    "license": item.get("license"),
                    "category": item.get("category"),
                },
            )
        )
    rows.sort(key=lambda row: row.rank)
    return rows


def _snapshot(
    *,
    id: str,
    name: str,
    category: str,
    description: str,
    source_url: str,
    methodology_url: str | None,
    authenticity: str,
    update_strategy: str,
    fetcher: Callable[[], list[BenchmarkRow]],
    min_rows: int = 3,
    score_range: tuple[float, float] | None = None,
) -> BenchmarkSnapshot:
    fetched_at = now_iso()
    try:
        rows = _fetch_with_retry(fetcher)
        warning = _validate_snapshot(rows, min_rows=min_rows, score_range=score_range)
        return BenchmarkSnapshot(
            id=id,
            name=name,
            category=category,
            description=description,
            source_url=source_url,
            methodology_url=methodology_url,
            authenticity=authenticity,
            update_strategy=update_strategy,
            fetched_at=fetched_at,
            status="ok",
            warning=warning,
            rows=rows[:30],
        )
    except Exception as exc:  # noqa: BLE001 - stored for dashboard diagnostics
        return BenchmarkSnapshot(
            id=id,
            name=name,
            category=category,
            description=description,
            source_url=source_url,
            methodology_url=methodology_url,
            authenticity=authenticity,
            update_strategy=update_strategy,
            fetched_at=fetched_at,
            status="failed",
            rows=[],
            error=f"{type(exc).__name__}: {exc}",
        )


def collect_deepswe() -> BenchmarkSnapshot:
    def fetcher() -> list[BenchmarkRow]:
        rows = parse_deepswe_html(fetch_text("https://deepswe.datacurve.ai/"))
        if rows:
            return rows
        return parse_benchlm_next_data(fetch_text("https://benchlm.ai/benchmarks/deepSwe"))

    return _snapshot(
        id="deep_swe",
        name="DeepSWE",
        category="Coding agents",
        description="Long-horizon software engineering tasks written from scratch across active repositories.",
        source_url="https://deepswe.datacurve.ai/",
        methodology_url="https://deepswe.datacurve.ai/blog",
        authenticity="Original tasks plus behavioural verifiers make it a stronger signal than saturated issue-derived coding benchmarks, though public tasks can still become contaminated over time.",
        update_strategy="Fetch official Datacurve page and parse embedded leaderboard rows; fall back to BenchLM DeepSWE mirror if the official payload shape changes.",
        fetcher=fetcher,
        min_rows=5,
        score_range=(0.0, 100.0),
    )


def collect_terminal_bench() -> BenchmarkSnapshot:
    version_url = "https://www.tbench.ai/leaderboard/terminal-bench/2.1"

    def fetcher() -> list[BenchmarkRow]:
        rows = parse_terminal_html(fetch_text(version_url))
        if len(rows) < 5:
            rows = parse_terminal_html(fetch_text("https://www.tbench.ai/leaderboard/terminal-bench/2.0"))
            for row in rows:
                row.metadata["fallback_version"] = "2.0"
        return rows

    return _snapshot(
        id="terminal_bench",
        name="Terminal-Bench",
        category="Terminal/tool use",
        description="Agent performance on real terminal tasks: shell commands, debugging, coding, sysadmin, and multi-step CLI workflows.",
        source_url=version_url,
        methodology_url="https://www.tbench.ai/",
        authenticity="Real terminal execution is close to practical agent work. Public tasks and scaffold differences still create overfitting and comparability risks.",
        update_strategy="Parse the official Terminal-Bench 2.1 table; if it becomes sparse/unavailable, use the official 2.0 table as a labelled fallback.",
        fetcher=fetcher,
        min_rows=3,
        score_range=(0.0, 100.0),
    )


def collect_gdpval_aa() -> BenchmarkSnapshot:
    def fetcher() -> list[BenchmarkRow]:
        try:
            rows = parse_aa_gdpval_table(fetch_text("https://artificialanalysis.ai/evaluations/gdpval-aa"))
            if rows:
                return rows
        except Exception:  # noqa: BLE001
            pass
        return parse_benchlm_next_data(
            fetch_text("https://benchlm.ai/benchmarks/gdpvalAa"),
            score_unit="Elo rating",
        )

    return _snapshot(
        id="gdpval_aa",
        name="GDPval-AA",
        category="Agentic professional work",
        description="1,320 real professional deliverables across 44 occupations (legal briefs, engineering plans, nursing care plans, customer support). Head-to-head Elo scored by Artificial Analysis.",
        source_url="https://artificialanalysis.ai/evaluations/gdpval-aa",
        methodology_url="https://openai.com/index/gdpval/",
        authenticity="Tasks are actual professional deliverables, not synthetic QA, making direct optimisation harder. Elo head-to-head scoring by independent evaluator Artificial Analysis avoids provider self-reporting. Tasks created by OpenAI in collaboration with industry professionals.",
        update_strategy="Parse Artificial Analysis HTML table directly (they run the evaluation); fall back to BenchLM mirror if the official page structure changes.",
        fetcher=fetcher,
        min_rows=10,
        score_range=(500.0, 2500.0),
    )


def collect_browsecomp() -> BenchmarkSnapshot:
    return _snapshot(
        id="browsecomp",
        name="BrowseComp",
        category="Browser/web research",
        description="Hard web-browsing questions requiring persistent search, source inspection, and concise verified answers.",
        source_url="https://benchlm.ai/benchmarks/browseComp",
        methodology_url="https://openai.com/index/browsecomp/",
        authenticity="Closer to real research than static QA, but the fixed 1,266-question corpus is now public and the benchmark is approaching saturation for frontier models (~88–90% cluster). Use as a supporting signal rather than a primary differentiator.",
        update_strategy="Parse BenchLM's benchmark-specific Next.js data for BrowseComp public leaderboard rows; link back to OpenAI's original benchmark methodology.",
        fetcher=lambda: parse_benchlm_next_data(fetch_text("https://benchlm.ai/benchmarks/browseComp")),
        min_rows=5,
        score_range=(0.0, 100.0),
    )


def collect_osworld() -> BenchmarkSnapshot:
    def fetcher() -> list[BenchmarkRow]:
        try:
            rows = parse_osworld_workbook(fetch_bytes("https://os-world.github.io/static/data/osworld_verified_results.xlsx"))
            if rows:
                return rows
        except Exception:  # noqa: BLE001
            pass
        return parse_benchlm_next_data(
            fetch_text("https://benchlm.ai/benchmarks/osWorldVerified"),
            score_unit="% success rate",
        )

    return _snapshot(
        id="osworld_verified",
        name="OSWorld-Verified",
        category="Computer use",
        description="Open-ended GUI/computer-use tasks in real desktop environments across apps and workflows.",
        source_url="https://os-world.github.io/",
        methodology_url="https://github.com/xlang-ai/OSWorld",
        authenticity="Execution in real desktop environments makes simple memorisation less useful; environment drift and agent-scaffold differences still matter.",
        update_strategy="Download official OSWorld-Verified XLSX from os-world.github.io; fall back to BenchLM mirror if the file path changes.",
        fetcher=fetcher,
        min_rows=5,
        score_range=(0.0, 100.0),
    )


def collect_longbench_v2() -> BenchmarkSnapshot:
    def fetcher() -> list[BenchmarkRow]:
        rows = parse_longbench_html(fetch_text("https://longbench2.github.io/"))
        if rows:
            return rows
        return parse_benchlm_next_data(fetch_text("https://benchlm.ai/benchmarks/longBenchV2"))

    return _snapshot(
        id="longbench_v2",
        name="LongBench v2",
        category="Long-context reasoning",
        description="Long-context reasoning and retrieval benchmark testing whether models can use extended context windows across diverse task types. Sourced directly from the official leaderboard.",
        source_url="https://longbench2.github.io/",
        methodology_url="https://arxiv.org/abs/2412.15204",
        authenticity="Official leaderboard with direct lab submissions. Tests real context utilisation, not just window size. Model set reflects academic submission cadence — frontier labs submit when ready, not on a fixed schedule.",
        update_strategy="Parse official LongBench v2 HTML table from longbench2.github.io; fall back to BenchLM mirror if official site structure changes.",
        fetcher=fetcher,
        min_rows=5,
        score_range=(0.0, 100.0),
    )


_LMARENA_HF_BASE = "https://datasets-server.huggingface.co/rows?dataset=lmarena-ai%2Fleaderboard-dataset&split=latest&offset=0&length=50"
# HF dataset configs in preference order; the first one that returns rows wins.
_LMARENA_CONFIGS = ("text", "text_style_control")


def collect_lmarena_text() -> BenchmarkSnapshot:
    def fetcher() -> list[BenchmarkRow]:
        last_exc: Exception | None = None
        for config in _LMARENA_CONFIGS:
            try:
                url = f"{_LMARENA_HF_BASE}&config={config}"
                response = requests.get(url, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
                response.raise_for_status()
                rows = parse_lmarena_rows(response.json())
                if rows:
                    return rows
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
        if last_exc is not None:
            raise last_exc
        return []

    return _snapshot(
        id="lmarena_text_style",
        name="LMArena Text",
        category="Writing preference",
        description="Human-preference arena ratings for text models, used here as a pragmatic proxy for prose and style quality.",
        source_url="https://huggingface.co/datasets/lmarena-ai/leaderboard-dataset",
        methodology_url="https://arena.ai/",
        authenticity="Live human preference battles are harder to optimise directly than static tests, but they reflect popularity/style preferences rather than academic correctness.",
        update_strategy="Fetch the public Hugging Face dataset-server rows for the LMArena text latest split (falls back to text_style_control if the primary config is unavailable).",
        fetcher=fetcher,
        min_rows=10,
        score_range=(900.0, 2200.0),
    )


def collect_aa_intelligence_index() -> BenchmarkSnapshot:
    return _snapshot(
        id="aa_intelligence_index",
        name="AA Intelligence Index",
        category="Composite intelligence",
        description=(
            "Artificial Analysis Intelligence Index — a composite of 9 evaluations run independently by "
            "Artificial Analysis rather than self-reported by labs: GDPval-AA (professional deliverables), "
            "τ³-Banking (agentic tool use), Terminal-Bench (terminal tasks), SciCode (scientific coding), "
            "Humanity's Last Exam and GPQA Diamond (frontier knowledge), CritPt (physics research), "
            "AA-Omniscience (knowledge/hallucination) and AA-LCR (long-context retrieval). Higher is better."
        ),
        source_url="https://artificialanalysis.ai/",
        methodology_url="https://artificialanalysis.ai/methodology/intelligence-benchmarking",
        authenticity=(
            "Every component evaluation is run independently by Artificial Analysis on its own harnesses, "
            "not self-reported by providers, and the mix spans reasoning, knowledge, agentic, coding and "
            "long-context tasks, so a model cannot top the index on a single narrow strength. It is still "
            "a weighted composite: the choice and weighting of components is Artificial Analysis' editorial "
            "decision, and the index version evolves over time (v4.1 at integration). Use as a headline "
            "general-capability signal alongside — not instead of — the task-specific benchmarks above."
        ),
        update_strategy=(
            "Parse the schema.org Dataset JSON-LD blocks embedded in the Artificial Analysis homepage, "
            "matching the exact dataset name 'Artificial Analysis Intelligence Index' (score key "
            "'intelligenceIndex'). If AA renames the headline block, fall back to the shorter 'Intelligence' "
            "dataset (score key 'artificialAnalysisIntelligenceIndex') from the same page."
        ),
        fetcher=lambda: parse_aa_intelligence_index_ldjson(fetch_text("https://artificialanalysis.ai/")),
        min_rows=10,
        score_range=(0.0, 100.0),
    )


# ── HF Open LLM Leaderboard 2 ─────────────────────────────────────────────────

_HF_LLB2_DATASET = "open-llm-leaderboard%2Fcontents"
_HF_LLB2_PARQUET_API = f"https://datasets-server.huggingface.co/parquet?dataset={_HF_LLB2_DATASET}"

# Columns we expect; validated on load so schema drift causes a clear failure.
_HF_LLB2_REQUIRED_COLS = {
    "#Params (B)",
    "Average ⬆️",
    "fullname",
    "Flagged",
    "Merged",
    "Available on the hub",
    "Weight type",
    "Type",
    "Hub ❤️",
}
# Minimum Hub likes required to appear in the listing.
# Benchmark-gaming fine-tunes have 0–3 likes; real lab models have hundreds.
# 50 is the sweet spot: keeps Falcon3, drops Tsunami/T.E-8.1 etc.
_HF_LLB2_MIN_LIKES = 50
_HF_LLB2_BENCH_COLS = ["IFEval", "BBH", "MATH Lvl 5", "GPQA", "MUSR", "MMLU-PRO"]

# ── BigCodeBench ───────────────────────────────────────────────────────────────

_HF_BCB_DATASET = "bigcode%2Fbigcodebench-results"
_HF_BCB_PARQUET_API = f"https://datasets-server.huggingface.co/parquet?dataset={_HF_BCB_DATASET}"
_HF_BCB_REQUIRED_COLS = {"model", "size", "instruct"}


def _fetch_hf_llb2_parquet() -> bytes:
    """Discover the current Parquet URL via the datasets-server API and download it."""
    meta = requests.get(_HF_LLB2_PARQUET_API, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    meta.raise_for_status()
    files = meta.json().get("parquet_files", [])
    if not files:
        raise ValueError("No Parquet files listed by datasets-server for open-llm-leaderboard/contents")
    url = files[0]["url"]
    resp = requests.get(url, timeout=120, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    return resp.content


def _fetch_bigcodebench_parquet() -> bytes:
    """Discover the current BigCodeBench Parquet URL via datasets-server API and download it."""
    meta = requests.get(_HF_BCB_PARQUET_API, timeout=TIMEOUT, headers={"User-Agent": USER_AGENT})
    meta.raise_for_status()
    files = meta.json().get("parquet_files", [])
    if not files:
        raise ValueError("No Parquet files listed by datasets-server for bigcode/bigcodebench-results")
    url = files[0]["url"]
    resp = requests.get(url, timeout=60, headers={"User-Agent": USER_AGENT})
    resp.raise_for_status()
    return resp.content


def parse_bigcodebench_parquet(
    content: bytes,
    params_min: float,
    params_max: float,
    limit: int = 10,
) -> list[BenchmarkRow]:
    """Parse BigCodeBench results Parquet and return top-N instruct models in a param range."""
    df = pd.read_parquet(io.BytesIO(content))
    missing = _HF_BCB_REQUIRED_COLS - set(df.columns)
    if missing:
        raise ValueError(f"BigCodeBench Parquet schema changed — missing columns: {missing}")

    df = df[
        df["size"].notna()
        & (df["size"] > params_min)
        & (df["size"] <= params_max)
        & df["instruct"].notna()
    ]
    df = df.sort_values("instruct", ascending=False).head(limit).reset_index(drop=True)

    rows: list[BenchmarkRow] = []
    for _, row in df.iterrows():
        model_name = clean_text(row.get("model", ""))
        if not model_name:
            continue
        instruct_score = as_float(row.get("instruct"))
        if instruct_score is None:
            continue
        meta: dict[str, Any] = {}
        complete_score = as_float(row.get("complete"))
        if complete_score is not None:
            meta["bcb_complete"] = round(complete_score, 1)
        params = as_float(row.get("size"))
        if params is not None:
            meta["params_b"] = round(params, 1)
        if row.get("moe"):
            meta["moe"] = True
        rows.append(
            BenchmarkRow(
                rank=len(rows) + 1,
                model=model_name,
                organization=infer_org(model_name),
                score=round(instruct_score, 1),
                score_unit="% instruct",
                date=clean_text(row.get("date", "")) or None,
                metadata=meta,
            )
        )
    return rows


def parse_hf_llm_leaderboard_parquet(
    content: bytes,
    params_min: float,
    params_max: float,
    limit: int = 10,
) -> list[BenchmarkRow]:
    """Parse HF Open LLM Leaderboard 2 Parquet bytes and return top-N models in a param range."""
    df = pd.read_parquet(io.BytesIO(content))
    missing = _HF_LLB2_REQUIRED_COLS - set(df.columns)
    if missing:
        raise ValueError(f"HF LLB2 Parquet schema changed — missing columns: {missing}")

    # Exclude merge models: both the Merged flag AND the Type label (inconsistently applied).
    is_merge_type = df["Type"].str.contains("merge", case=False, na=False)
    # Exclude quantized repacks — bnb-4bit/bnb-8bit/gguf variants report inflated param counts.
    _QUANT_PATTERNS = ("bnb-4bit", "bnb-8bit", "-gguf", ".gguf", "-4bit", "-8bit")
    is_quantized = df["fullname"].str.lower().str.contains(
        "|".join(_QUANT_PATTERNS), regex=True, na=False
    )
    df = df[
        (df["#Params (B)"] > params_min)
        & (df["#Params (B)"] <= params_max)
        & (df["Flagged"] == False)  # noqa: E712
        & (df["Merged"] == False)  # noqa: E712
        & (~is_merge_type)
        & (~is_quantized)
        & (df["Available on the hub"] == True)  # noqa: E712
        & (df["Weight type"] == "Original")
        & (df["Hub ❤️"] >= _HF_LLB2_MIN_LIKES)
    ]
    df = df.sort_values("Average ⬆️", ascending=False).head(limit).reset_index(drop=True)

    rows: list[BenchmarkRow] = []
    for idx, row in df.iterrows():
        fullname = clean_text(row.get("fullname", ""))
        org = fullname.split("/")[0] if "/" in fullname else None
        avg = as_float(row.get("Average ⬆️"))
        if avg is None:
            continue
        bench_meta: dict[str, Any] = {}
        for col in _HF_LLB2_BENCH_COLS:
            val = as_float(row.get(col))
            if val is not None:
                bench_meta[col] = round(val, 1)
        bench_meta["params_b"] = round(float(row["#Params (B)"]), 1)
        bench_meta["type"] = clean_text(row.get("Type", ""))
        upload_date = clean_text(row.get("Upload To Hub Date", "")) or None
        rows.append(
            BenchmarkRow(
                rank=len(rows) + 1,
                model=fullname,
                organization=normalise_org(org) or infer_org(fullname),
                score=round(avg, 2),
                score_unit="avg %",
                date=upload_date,
                metadata=bench_meta,
            )
        )
    return rows


def _collect_hf_local(
    *,
    id: str,
    name: str,
    params_min: float,
    params_max: float,
    description: str,
) -> BenchmarkSnapshot:
    def fetcher() -> list[BenchmarkRow]:
        return parse_hf_llm_leaderboard_parquet(
            _fetch_hf_llb2_parquet(), params_min=params_min, params_max=params_max
        )

    return _snapshot(
        id=id,
        name=name,
        category="Local inference",
        description=description,
        source_url="https://huggingface.co/spaces/open-llm-leaderboard/open_llm_leaderboard",
        methodology_url="https://huggingface.co/docs/leaderboards/en/open_llm_leaderboard/about",
        authenticity=(
            "HF Open LLB2 uses six standardised benchmarks (IFEval, BBH, MATH Lvl 5, GPQA, MUSR, "
            "MMLU-PRO) run by a third party on a reproducible harness — more trustworthy than "
            "self-reported lab numbers. Restricted to models with ≥50 Hub likes, which cleanly "
            "separates real lab models from benchmark-gaming fine-tunes (the latter have 0–3 likes). "
            "Very new releases may not yet be submitted; they appear once the community or lab "
            "submits them to the leaderboard."
        ),
        update_strategy=(
            "Download official HF LLB2 Parquet via datasets-server /parquet API (URL discovered "
            "dynamically each run). Filter: param range, Hub likes ≥ 50, Flagged=False, "
            "Merged=False, Type not containing 'merge', Weight type=Original, no quantized-repack "
            "name suffixes (bnb-4bit, gguf, etc). Sort by composite Average, top 10."
        ),
        fetcher=fetcher,
        min_rows=5,
        score_range=(0.0, 100.0),
    )


def collect_hf_local_under10b() -> BenchmarkSnapshot:
    return _collect_hf_local(
        id="hf_local_under10b",
        name="Local models < 10B",
        params_min=0.0,
        params_max=10.0,
        description=(
            "Best general-capability open-weight models under 10 billion parameters, ranked by the "
            "HF Open LLM Leaderboard 2 composite score. The composite averages six standardised "
            "benchmarks: IFEval (can the model follow precise, verifiable instructions?), BBH "
            "(multi-step reasoning across 23 challenging BIG-Bench Hard tasks), MATH Lvl 5 "
            "(competition-level mathematics at AMC/AIME difficulty), GPQA (graduate-level science "
            "questions written by PhD researchers), MUSR (multi-step soft reasoning over long "
            "narratives), and MMLU-PRO (expert professional knowledge across 14 domains). These "
            "rankings reflect general academic capability, not task-specific tuning — a high score "
            "means the model is broadly capable. Runnable on a laptop GPU or Apple Silicon. "
            "Rankings cover only models submitted to the leaderboard; very new releases may appear "
            "after community or lab submission."
        ),
    )


def collect_hf_local_10to20b() -> BenchmarkSnapshot:
    return _collect_hf_local(
        id="hf_local_10to20b",
        name="Local models 10–20B",
        params_min=10.0,
        params_max=20.0,
        description=(
            "Best general-capability open-weight models in the 10–20 billion parameter range, "
            "ranked by the HF Open LLM Leaderboard 2 composite score. The composite averages six "
            "standardised benchmarks: IFEval (can the model follow precise, verifiable "
            "instructions?), BBH (multi-step reasoning across 23 challenging BIG-Bench Hard tasks), "
            "MATH Lvl 5 (competition-level mathematics at AMC/AIME difficulty), GPQA "
            "(graduate-level science questions written by PhD researchers), MUSR (multi-step soft "
            "reasoning over long narratives), and MMLU-PRO (expert professional knowledge across 14 "
            "domains). These rankings reflect general academic capability, not task-specific tuning "
            "— a high score means the model is broadly capable. Runnable on a single consumer GPU "
            "with 16–24 GB VRAM. Rankings cover only models submitted to the leaderboard; very new "
            "releases may appear after community or lab submission."
        ),
    )


def _collect_bcb(
    *,
    id: str,
    name: str,
    params_min: float,
    params_max: float,
    description: str,
) -> BenchmarkSnapshot:
    def fetcher() -> list[BenchmarkRow]:
        return parse_bigcodebench_parquet(
            _fetch_bigcodebench_parquet(), params_min=params_min, params_max=params_max
        )

    return _snapshot(
        id=id,
        name=name,
        category="Local inference",
        description=description,
        source_url="https://huggingface.co/spaces/bigcode/bigcodebench-leaderboard",
        methodology_url="https://bigcodebench.github.io/",
        authenticity=(
            "BigCodeBench tasks are directly executable — each solution is verified by running it "
            "against real library calls, not text matching. Curated by the BigCode team "
            "(independent open-source researchers behind The Stack and StarCoder), not by model "
            "providers. Models are submitted by the community. The task corpus is static so "
            "contamination risk increases as newer models may have seen these tasks in training."
        ),
        update_strategy=(
            "Download BigCodeBench results Parquet via datasets-server /parquet API (URL discovered "
            "dynamically each run). Filter: parameter range, non-null instruct score. Sort by "
            "instruct score descending, top 10."
        ),
        fetcher=fetcher,
        min_rows=3,
        score_range=(0.0, 100.0),
    )


def collect_bigcodebench_under10b() -> BenchmarkSnapshot:
    return _collect_bcb(
        id="bigcodebench_under10b",
        name="Local coding < 10B",
        params_min=0.0,
        params_max=10.0,
        description=(
            "Best coding-capable open-weight models under 10 billion parameters, ranked by "
            "BigCodeBench Instruct score. BigCodeBench tests 1,140 practical coding tasks that "
            "require calling real Python libraries (NumPy, pandas, requests, PIL, and 130+ others) "
            "— significantly harder than classic HumanEval because tasks demand multi-library "
            "orchestration, not just pattern completion. The Instruct split used for ranking "
            "evaluates models on natural-language coding instructions, mirroring how developers "
            "actually prompt models. The Complete split (docstring prompts) score is shown as a "
            "secondary note. Runnable on a laptop GPU or Apple Silicon; coding-focused scores "
            "complement the general-capability HF LLB2 ranking."
        ),
    )


def collect_bigcodebench_10to20b() -> BenchmarkSnapshot:
    return _collect_bcb(
        id="bigcodebench_10to20b",
        name="Local coding 10–20B",
        params_min=10.0,
        params_max=20.0,
        description=(
            "Best coding-capable open-weight models in the 10–20 billion parameter range, ranked "
            "by BigCodeBench Instruct score. BigCodeBench tests 1,140 practical coding tasks that "
            "require calling real Python libraries (NumPy, pandas, requests, PIL, and 130+ others) "
            "— significantly harder than classic HumanEval because tasks demand multi-library "
            "orchestration, not just pattern completion. The Instruct split used for ranking "
            "evaluates models on natural-language coding instructions, mirroring how developers "
            "actually prompt models. The Complete split (docstring prompts) score is shown as a "
            "secondary note. Runnable on a single consumer GPU with 16–24 GB VRAM; coding-focused "
            "scores complement the general-capability HF LLB2 ranking."
        ),
    )


def collect_all() -> list[BenchmarkSnapshot]:
    collectors = [
        collect_deepswe,
        collect_gdpval_aa,
        collect_terminal_bench,
        collect_browsecomp,
        collect_osworld,
        collect_longbench_v2,
        collect_lmarena_text,
        collect_aa_intelligence_index,
        collect_hf_local_under10b,
        collect_hf_local_10to20b,
        collect_bigcodebench_under10b,
        collect_bigcodebench_10to20b,
    ]
    return [collector() for collector in collectors]
